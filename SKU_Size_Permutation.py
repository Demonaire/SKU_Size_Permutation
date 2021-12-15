'''
generates permutation of defined SKUs with provided sizes
Keeps original file format intact
Note: This task can easily be performed in excel. Intended use is for large files where its not feasible to work in excel
    for such data

'''

## Library Imports

from openpyxl import load_workbook
import openpyxl
from openpyxl import Workbook
from copy import copy
import shutil
import os


#Copy cell format from original cell to target cell along with defined value

def copy_format(original_cell,new_sheet,column,row,value):
        
        new_cell=new_sheet.cell(column=column,row=row,value=value)
        if cell.has_style:
            new_cell.font = copy(original_cell.font)
            new_cell.border = copy(original_cell.border)
            new_cell.fill = copy(original_cell.fill)
            new_cell.number_format = copy(original_cell.number_format)
            new_cell.protection = copy(original_cell.protection)
            new_cell.alignment = copy(original_cell.alignment)

#File Names
file_name='B1-Rings(Pending) bijans draft.xlsx'
file_new=file_name.replace('.xlsx','_mod.xlsx')

#Load original File to openpyxl
workbook = load_workbook(filename=file_name,data_only=True)

sheet = workbook['Jewelry']

#Sizes to be added for each SKU
sizes=[3.5, 3.75, 4, 4.25, 4.5, 4.75, 5, 5.25, 5.5, 5.75, 6, 6.25, 
       6.5, 6.75, 7, 7.25, 7.5, 7.75, 8, 8.25, 8.5, 8.75, 9, 9.25, 
       9.5, 9.75, 10, 10.25, 10.5, 10.75, 11]

#Make copy of original file
shutil.os.chdir(os.getcwd())
original = file_name
target =file_new

shutil.copyfile(original, target)

#Load new File in openpyxl
new_workbook=load_workbook(filename=file_new,data_only=True)
new_sheet=new_workbook['Jewelry']

offset=7

#iterate rows
for i,row in enumerate(sheet.iter_rows(),start=1):
    
    #ignore heading rows
    if i<offset:
        continue
        
    
    else:
        #iterate columns
        for j,cell in enumerate(row,start=1):
            
            new_row=offset+(i-offset)*len(sizes)
            
            #one entry for each size in list for each SKU
            for k,size in enumerate(sizes):
                
                #modify SKU along with size code
                if j==4:
                    copy_format(cell,new_sheet,column=j,row=new_row+k,value=(cell.value+'/SZ/'+str(size)))
                
                #input size in defined column
                elif j==22:
                    copy_format(cell,new_sheet,column=j,row=new_row+k,value=size)
                    
               #input size in defined column                
                elif j==36:
                    copy_format(cell,new_sheet,column=j,row=new_row+k,value=size)
                    
                #copy format along with value from original cell
                else:
                    copy_format(cell,new_sheet,column=j,row=new_row+k,value=cell.value)

#export modified file to system                    
new_workbook.save(file_new)